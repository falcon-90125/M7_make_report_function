from openpyxl import load_workbook
import openpyxl.styles.numbers

from data_visits import data_visits
from popular_goods import popular_goods
from genders_goods import genders_goods

# 7 самых популярных браузеров по посещаемости и товаров
log_file_name = 'logs.xlsx'
report_template_file_name = 'report_template.xlsx'
report_output_file_name = 'report_template_output.xlsx'

def make_report(log_file_name, report_template_file_name, report_output_file_name):
    # Посетители веб-сайта
    # browser_data - Данные по посещениям браузеров по месяцам
    # all_visits_pop_browser_months_dict - Словарь суммарных значений продаж по месяцам по всем популярным браузерам
    browser_data, all_visits_pop_browser_months_dict = data_visits(log_file_name)

    # Популярные товары
    # sales_goods_data - Данные по продажам товаров по месяцам
    # all_sales_pop_goods_months_dict - Словарь суммарных значений продаж по месяцам по всем популярным товарам
    sales_goods_data, all_sales_pop_goods_months_dict = popular_goods(log_file_name)

    # Предпочтения
    # Наиболее популярные товары: goods_list[0]mens_goods, goods_list[2]woomens_goods
    # Наименее популярные товары: goods_list[1]mens_goods, goods_list[3]woomens_goods
    genders_goods_list = genders_goods(log_file_name)

    # Открываем файл шаблона отчета report_template_file_name.xlsx
    wb = load_workbook(filename=report_template_file_name, data_only=True)
    ws = wb.active #1й лист [wb.sheetnames[0]]

    # Выполняем запись данных в объект wb

    # Посетители веб-сайта
    for i, val in enumerate(browser_data.keys()): # Задаём счётчик и наименования популярных браузеров browser_pop_list
        cell_i = ws.cell(row=5+i, column=1) # Проходим по ячейкам наименований популярных браузеров
        cell_i.value = val # Записываем наименования браузеров в 1ю колонку "Браузер"
        for month in range(len(all_visits_pop_browser_months_dict)): # Задаём месяцы посещений браузера
            cell_month = ws.cell(row=5+i, column=2+month) # Проходим по ячейкам для данных посещений по месяцам
            cell_month.value = browser_data[val][month+1] # Записываем значение посещения в ячейку
            cell_all_month = ws.cell(row=12, column=2+month) # Проходим по ячейкам для суммарных значений по месяцам всех браузеров
            cell_all_month.number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[0]
            cell_all_month.value = all_visits_pop_browser_months_dict[month+1] # Записываем итоговое значение посещения в ячейку

    # Популярные товары
    for i, val in enumerate(sales_goods_data.keys()):
        cell_i = ws.cell(row=19+i, column=1)
        cell_i.value = val
        for month in range(len(all_sales_pop_goods_months_dict)): # Задаём месяцы посещений браузера
            cell_month = ws.cell(row=19+i, column=2+month) # Проходим по ячейкам для данных посещений по месяцам
            cell_month.value = sales_goods_data[val][month+1] # Записываем значение посещения в ячейку
            cell_all_month = ws.cell(row=26, column=2+month) # Проходим по ячейкам для суммарных значений по месяцам всех браузеров
            cell_all_month.number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[0]
            cell_all_month.value = all_sales_pop_goods_months_dict[month+1] # Записываем итоговое значение посещения в ячейку

    # Предпочтения
    ws['B31'] = genders_goods_list[0] #pop_mens_goods
    ws['B32'] = genders_goods_list[2] #pop_woomens_goods
    ws['B33'] = genders_goods_list[1] #no_pop_mens_goods
    ws['B34'] = genders_goods_list[3] #no_pop_woomens_goods

    # Сохраняем файл-отчет
    wb.save(report_output_file_name)

make_report(log_file_name, report_template_file_name, report_output_file_name)
